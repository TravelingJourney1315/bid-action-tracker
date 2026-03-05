import streamlit as st
import openpyxl
from openpyxl import load_workbook
import datetime
import io
import os
import pathlib

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Bid Action Tracker",
    page_icon="📋",
    layout="wide",
)

# ── Color constants ────────────────────────────────────────────────────────────
ORANGE = "FC4C01"
DARK   = "33353B"
PURPLE = "542EFD"
WHITE  = "FFFFFF"
LIGHT  = "F0EEFF"

# ── Custom CSS ─────────────────────────────────────────────────────────────────
st.markdown(f"""
<style>
  @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;600;700;800&display=swap');

  html, body, [class*="css"] {{
    font-family: 'DM Sans', sans-serif;
    color: #{DARK};
  }}

  /* Header bar */
  .header-bar {{
    background: #{DARK};
    border-radius: 14px;
    padding: 22px 30px;
    display: flex;
    align-items: center;
    gap: 16px;
    margin-bottom: 28px;
  }}
  .header-icon {{
    width: 48px; height: 48px;
    background: linear-gradient(135deg, #{PURPLE}, #{ORANGE});
    border-radius: 12px;
    display: flex; align-items: center; justify-content: center;
    font-size: 24px;
  }}
  .header-title {{ color: white; font-size: 22px; font-weight: 800; margin: 0; }}
  .header-sub   {{ color: #aaa; font-size: 13px; margin: 0; }}

  /* Section headers */
  .section-header {{
    border-left: 4px solid #{PURPLE};
    padding-left: 12px;
    margin: 8px 0 18px;
    font-size: 15px;
    font-weight: 700;
    color: #{PURPLE};
  }}

  /* Template toggle buttons */
  div.stButton > button {{
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 700 !important;
    border-radius: 30px !important;
    border: 2px solid #{PURPLE} !important;
    transition: all 0.2s !important;
    white-space: nowrap !important;
    overflow: hidden !important;
  }}

  /* Generate button */
  .gen-btn > button {{
    background: linear-gradient(135deg, #{PURPLE}, #8B5CF6) !important;
    color: white !important;
    border: none !important;
    padding: 14px 40px !important;
    font-size: 16px !important;
    font-weight: 800 !important;
    border-radius: 40px !important;
    box-shadow: 0 6px 24px rgba(84,46,253,0.35) !important;
  }}

  /* Card-like containers */
  .card {{
    background: white;
    border-radius: 14px;
    padding: 24px 28px;
    box-shadow: 0 2px 16px rgba(84,46,253,0.07);
    border: 1px solid #E2E0F0;
    margin-bottom: 20px;
  }}

  /* Task table headers */
  .task-header {{
    background: #{PURPLE};
    color: white;
    border-radius: 8px;
    padding: 10px 14px;
    font-weight: 700;
    font-size: 13px;
  }}

  /* Download button */
  .stDownloadButton > button {{
    background: linear-gradient(135deg, #{PURPLE}, #8B5CF6) !important;
    color: white !important;
    border: none !important;
    border-radius: 30px !important;
    font-weight: 700 !important;
    font-size: 15px !important;
    padding: 12px 36px !important;
    box-shadow: 0 6px 24px rgba(84,46,253,0.35) !important;
  }}

  /* Progress bar color */
  .stProgress > div > div > div > div {{
    background: linear-gradient(90deg, #{PURPLE}, #{ORANGE}) !important;
  }}

  /* Input borders */
  input[type="text"], input[type="date"] {{
    border: 1.5px solid #E2E0F0 !important;
    border-radius: 8px !important;
  }}
  input:focus {{
    border-color: #{PURPLE} !important;
  }}

  .success-box {{
    background: #edfaf2;
    border: 1px solid #1a7a4a;
    border-radius: 10px;
    padding: 14px 18px;
    color: #1a7a4a;
    font-weight: 600;
    font-size: 14px;
    margin-bottom: 16px;
  }}

  hr {{ border-color: #E2E0F0; }}
</style>
""", unsafe_allow_html=True)

# ── Header ─────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="header-bar">
  <div class="header-icon">📋</div>
  <div>
    <p class="header-title">Bid Action Tracker</p>
    <p class="header-sub">Auto-fill your RFP templates in seconds</p>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Task list ──────────────────────────────────────────────────────────────────
TASKS = [
    "RFP Timeline",
    "RFP draft",
    "Executive summary",
    "SCS Review",
    "First draft review",
    "Return edits",
    "Discuss draft updates",
    "Internal deadline",
    "Proofing",
    "Handoff",
    "Upload/Submit to client",
]

# ── Template Selection ─────────────────────────────────────────────────────────
st.markdown('<div class="section-header">Choose Template</div>', unsafe_allow_html=True)

col_r, col_e, col_spacer = st.columns([1.2, 1, 2.8])
with col_r:
    robust_btn = st.button("🗂 Robust Template", use_container_width=True)
with col_e:
    ez_btn = st.button("⚡ EZ Template", use_container_width=True)

if robust_btn:
    st.session_state["template"] = "robust"
if ez_btn:
    st.session_state["template"] = "ez"
if "template" not in st.session_state:
    st.session_state["template"] = "robust"

template = st.session_state["template"]

desc_map = {
    "robust": "**Robust** — Fills Calendar, Timeline, and RFP Project Plan tabs in the **Agile BAT Template**",
    "ez":     "**EZ** — Fills Calendar and Timeline tabs in the **Bid Tracker** template",
}
st.markdown(f"<small style='color:#8b8fa8'>{desc_map[template]}</small>", unsafe_allow_html=True)

# ── Embedded template loader ───────────────────────────────────────────────────
def load_template(template_type: str) -> bytes:
    """Load template from repo root, works locally and on Streamlit Cloud."""
    fname = "Agile_BAT_Template.xlsx" if template_type == "robust" else "Bid_Tracker.xlsx"

    candidates = [
        pathlib.Path(__file__).parent / fname,
        pathlib.Path(os.getcwd()) / fname,
    ]
    # Streamlit Cloud mounts repos under /mount/src/<repo-name>/
    mount = pathlib.Path("/mount/src")
    if mount.exists():
        candidates.append(mount / fname)
        for sub in mount.iterdir():
            if sub.is_dir():
                candidates.append(sub / fname)

    for path in candidates:
        if path.exists():
            return path.read_bytes()

    st.error(
        f"**Template not found: `{fname}`** \n\n"
        f"Commit `{fname}` to the root of your GitHub repo alongside `bat_app.py`."
    )
    st.stop()

st.markdown("---")

# ── Project Info ───────────────────────────────────────────────────────────────
st.markdown('<div class="section-header">🏗 Project Information</div>', unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1:
    project_name = st.text_input("Project Name *", placeholder="e.g. ACME Corp RFP 2026")
with col2:
    project_manager = st.text_input("Project Manager", placeholder="e.g. Journey Young")

st.markdown("---")

# ── Key Dates ──────────────────────────────────────────────────────────────────
st.markdown('<div class="section-header">📅 Key Dates</div>', unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1:
    start_date = st.date_input("Project Start Date *", value=None, key="start")
with col2:
    customer_deadline = st.date_input("Customer Deadline *", value=None, key="cust_dl")

col3, col4 = st.columns(2)
with col3:
    scs_deadline = st.date_input("Internal SCS Review Deadline", value=None, key="scs")
with col4:
    rfp_text_deadline = st.date_input("Internal RFP Text Deadline", value=None, key="rfp_text")

if template == "robust":
    col5, _ = st.columns(2)
    with col5:
        end_date = st.date_input("Project End Date", value=None, key="end",
                                  help="Used in Timeline & RFP Project Plan tabs")
else:
    end_date = None

st.markdown("---")

# ── Task Timeline ──────────────────────────────────────────────────────────────
st.markdown('<div class="section-header">⏱ Task Timeline Dates</div>', unsafe_allow_html=True)
st.caption("Set start and finish dates for each task in the Timeline tab.")

# Column headers
hcol1, hcol2, hcol3 = st.columns([3, 2, 2])
hcol1.markdown(f"<div style='background:#{PURPLE};color:white;border-radius:8px;padding:8px 12px;font-weight:700;font-size:13px;'>TASK</div>", unsafe_allow_html=True)
hcol2.markdown(f"<div style='background:#{PURPLE};color:white;border-radius:8px;padding:8px 12px;font-weight:700;font-size:13px;'>START DATE</div>", unsafe_allow_html=True)
hcol3.markdown(f"<div style='background:#{PURPLE};color:white;border-radius:8px;padding:8px 12px;font-weight:700;font-size:13px;'>FINISH DATE</div>", unsafe_allow_html=True)

task_dates = {}
for i, task in enumerate(TASKS):
    bg = "#faf9ff" if i % 2 == 0 else "#ffffff"
    c1, c2, c3 = st.columns([3, 2, 2])
    with c1:
        st.markdown(
            f"<div style='background:{bg};border:1px solid #E2E0F0;border-radius:8px;"
            f"padding:11px 14px;font-size:13px;font-weight:600;height:44px;"
            f"display:flex;align-items:center;'>{task}</div>",
            unsafe_allow_html=True,
        )
    with c2:
        ts = st.date_input("", value=None, key=f"task_start_{i}", label_visibility="collapsed")
    with c3:
        tf = st.date_input("", value=None, key=f"task_finish_{i}", label_visibility="collapsed")
    task_dates[task] = {"start": ts, "finish": tf}

st.markdown("---")

# ── Progress indicator ─────────────────────────────────────────────────────────
total_fields = 7 + len(TASKS) * 2
filled = sum([
    bool(project_name), bool(project_manager),
    bool(start_date), bool(customer_deadline),
    bool(scs_deadline), bool(rfp_text_deadline),
    bool(end_date),
])
for task in TASKS:
    filled += bool(task_dates[task]["start"]) + bool(task_dates[task]["finish"])

pct = int(filled / total_fields * 100)
st.caption(f"Form completion: **{pct}%**")
st.progress(pct / 100)

st.markdown("---")


# ── Calendar builder helper ────────────────────────────────────────────────────
def build_calendar_dates(start: datetime.date, end: datetime.date):
    """
    Returns list of (month_name, [list of Mon-Fri dates]) grouped by week-start (Monday).
    Each item: {'month_label': str or None, 'dates': [date or None, ...] x5 Mon-Fri}
    Month label is set on the first week of each new month appearing in that week.
    """
    if start is None or end is None:
        return []

    # Walk to the Monday of the week containing start
    day = start - datetime.timedelta(days=start.weekday())  # Monday of start week
    weeks = []
    current_month = None

    while day <= end + datetime.timedelta(days=6):
        week_dates = []
        week_months = set()
        for offset in range(5):  # Mon-Fri
            d = day + datetime.timedelta(days=offset)
            week_dates.append(d)
            week_months.add(d.month)

        # Month label: use the month of Monday, show label if changed
        week_month = day.month
        if week_month != current_month:
            month_label = day.strftime("%B")
            current_month = week_month
        else:
            month_label = None

        weeks.append({"month_label": month_label, "dates": week_dates})
        day += datetime.timedelta(days=7)

        if day > end + datetime.timedelta(days=6):
            break

    return weeks


# ── Excel fill function ────────────────────────────────────────────────────────
def fill_calendar_tab(ws, project_name, project_manager, scs_deadline,
                       rfp_text_deadline, customer_deadline, start_date):
    """Fill Calendar tab. Works for both templates (same structure)."""
    DATE_FMT = "MM/DD/YYYY"

    # Project Name: row 2 col 4
    if project_name:
        ws.cell(2, 4).value = project_name
    # Project Manager: row 3 col 4
    if project_manager:
        ws.cell(3, 4).value = project_manager
    # Deadline values: labels at col 8, values at col 10
    if scs_deadline:
        ws.cell(2, 10).value = scs_deadline
        ws.cell(2, 10).number_format = DATE_FMT
    if rfp_text_deadline:
        ws.cell(3, 10).value = rfp_text_deadline
        ws.cell(3, 10).number_format = DATE_FMT
    if customer_deadline:
        ws.cell(4, 10).value = customer_deadline
        ws.cell(4, 10).number_format = DATE_FMT

    # Rebuild calendar grid if we have both dates
    if start_date and customer_deadline:
        weeks = build_calendar_dates(start_date, customer_deadline)

        # Clear existing calendar data rows 7+
        for r in range(7, ws.max_row + 1):
            for c in [2, 4, 6, 8, 10]:
                ws.cell(r, c).value = None

        # Day-of-week cols: Mon=2, Tue=4, Wed=6, Thu=8, Fri=10
        DOW_COLS = [2, 4, 6, 8, 10]

        # Write header row (row 6) - day names (should already exist but ensure)
        day_names = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY"]
        for col, name in zip(DOW_COLS, day_names):
            if not ws.cell(6, col).value:
                ws.cell(6, col).value = name

        current_row = 7
        for week in weeks:
            if week["month_label"]:
                ws.cell(current_row, 2).value = week["month_label"]
                current_row += 1

            for col, d in zip(DOW_COLS, week["dates"]):
                ws.cell(current_row, col).value = d.day
            current_row += 4  # 3 empty content rows + next week row


def fill_timeline_tab(ws, project_name, project_manager, start_date, end_date, task_dates):
    """Fill Timeline tab. Works for both templates (same structure)."""
    DATE_FMT = "MM/DD/YYYY"

    if project_name:
        ws.cell(2, 4).value = project_name
    if project_manager:
        ws.cell(3, 4).value = project_manager
    if start_date:
        ws.cell(2, 6).value = start_date
        ws.cell(2, 6).number_format = DATE_FMT
    if end_date:
        ws.cell(2, 7).value = end_date
        ws.cell(2, 7).number_format = DATE_FMT

    # Task rows start at row 6 (0-indexed from TASKS list)
    for i, task in enumerate(TASKS):
        row = 6 + i
        td = task_dates.get(task, {})
        if td.get("start"):
            ws.cell(row, 4).value = td["start"]
            ws.cell(row, 4).number_format = DATE_FMT
        if td.get("finish"):
            ws.cell(row, 5).value = td["finish"]
            ws.cell(row, 5).number_format = DATE_FMT


def fill_project_plan_tab(ws, project_name, project_manager, start_date, end_date):
    """Fill RFP Project Plan tab (Robust only)."""
    DATE_FMT = "MM/DD/YYYY"
    if project_name:
        ws.cell(2, 4).value = project_name
    if project_manager:
        ws.cell(3, 4).value = project_manager
    if start_date:
        ws.cell(2, 6).value = start_date
        ws.cell(2, 6).number_format = DATE_FMT
    if end_date:
        ws.cell(2, 7).value = end_date
        ws.cell(2, 7).number_format = DATE_FMT


def generate_filled_workbook(template_bytes, template_type, project_name, project_manager,
                              scs_deadline, rfp_text_deadline, customer_deadline,
                              start_date, end_date, task_dates):
    wb = load_workbook(filename=io.BytesIO(template_bytes))

    if template_type == "robust":
        # Calendar tab
        ws_cal = wb["Calendar"]
        fill_calendar_tab(ws_cal, project_name, project_manager, scs_deadline,
                          rfp_text_deadline, customer_deadline, start_date)
        # Timeline tab
        ws_tl = wb["Timeline"]
        fill_timeline_tab(ws_tl, project_name, project_manager, start_date, end_date, task_dates)
        # RFP Project Plan tab
        ws_pp = wb["RFP Project Plan"]
        fill_project_plan_tab(ws_pp, project_name, project_manager, start_date, end_date)

    else:  # ez / Bid Tracker
        # Calendar tab (has a leading space in name)
        cal_name = " Calendar" if " Calendar" in wb.sheetnames else "Calendar"
        ws_cal = wb[cal_name]
        fill_calendar_tab(ws_cal, project_name, project_manager, scs_deadline,
                          rfp_text_deadline, customer_deadline, start_date)
        # Timeline tab
        ws_tl = wb["Timeline"]
        fill_timeline_tab(ws_tl, project_name, project_manager, start_date, end_date, task_dates)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# ── Generate & Download ────────────────────────────────────────────────────────
if not project_name:
    st.warning("Please enter a **Project Name** to continue.", icon="⚠️")
elif not start_date or not customer_deadline:
    st.warning("Please set both **Start Date** and **Customer Deadline** to continue.", icon="⚠️")
else:
    st.markdown(
        '<div style="text-align:center;margin:20px 0;">',
        unsafe_allow_html=True,
    )

    try:
        template_bytes = load_template(template)
        output_bytes = generate_filled_workbook(
            template_bytes=template_bytes,
            template_type=template,
            project_name=project_name,
            project_manager=project_manager,
            scs_deadline=scs_deadline,
            rfp_text_deadline=rfp_text_deadline,
            customer_deadline=customer_deadline,
            start_date=start_date,
            end_date=end_date,
            task_dates=task_dates,
        )

        fname = (
            f"{project_name.replace(' ', '_')}_Agile_BAT_filled.xlsx"
            if template == "robust"
            else f"{project_name.replace(' ', '_')}_Bid_Tracker_filled.xlsx"
        )

        st.success("✅ Template filled successfully! Click below to download.", icon="🎉")

        col_dl, _, _ = st.columns([1, 2, 2])
        with col_dl:
            st.download_button(
                label="⬇️  Download Filled Template",
                data=output_bytes,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        # Summary of what was filled
        st.markdown("#### 📋 What was filled:")
        filled_items = []
        if project_name:     filled_items.append(f"✅ Project Name: **{project_name}**")
        if project_manager:  filled_items.append(f"✅ Project Manager: **{project_manager}**")
        if start_date:       filled_items.append(f"✅ Start Date: **{start_date.strftime('%b %d, %Y')}**")
        if customer_deadline: filled_items.append(f"✅ Customer Deadline: **{customer_deadline.strftime('%b %d, %Y')}**")
        if scs_deadline:     filled_items.append(f"✅ SCS Review Deadline: **{scs_deadline.strftime('%b %d, %Y')}**")
        if rfp_text_deadline: filled_items.append(f"✅ RFP Text Deadline: **{rfp_text_deadline.strftime('%b %d, %Y')}**")
        if end_date and template == "robust":
            filled_items.append(f"✅ End Date: **{end_date.strftime('%b %d, %Y')}**")
        task_filled = sum(1 for t in TASKS if task_dates[t]["start"] or task_dates[t]["finish"])
        if task_filled:
            filled_items.append(f"✅ Task dates: **{task_filled} of {len(TASKS)} tasks** have dates set")

        tabs_filled = ["Calendar", "Timeline"]
        if template == "robust":
            tabs_filled.append("RFP Project Plan")
        filled_items.append(f"✅ Tabs updated: **{', '.join(tabs_filled)}**")
        filled_items.append(f"✅ Calendar rebuilt from **{start_date.strftime('%b %d')}** to **{customer_deadline.strftime('%b %d, %Y')}**")

        cols = st.columns(2)
        for i, item in enumerate(filled_items):
            cols[i % 2].markdown(item)

    except Exception as e:
        st.error(f"❌ Error generating file: {e}", icon="🚨")
        st.exception(e)

    st.markdown("</div>", unsafe_allow_html=True)

# ── Footer ─────────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown(
    f"<div style='text-align:center;color:#aaa;font-size:12px;padding:8px'>"
    f"Bid Action Tracker &nbsp;•&nbsp; Powered by openpyxl &nbsp;•&nbsp; "
    f"<span style='color:#{PURPLE}'>Robust</span> & "
    f"<span style='color:#{ORANGE}'>EZ</span> templates supported"
    f"</div>",
    unsafe_allow_html=True,
)
